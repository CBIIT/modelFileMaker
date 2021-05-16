#! /usr/bin/env python
#A script to create model description files in the YAML format.
from openpyxl import load_workbook, Workbook
from openpyxl.cell.read_only import EmptyCell
from collections import defaultdict, namedtuple
import sys
import re
import argparse

###################################################################
#Initialize space and separator variables for use in all functions.
#Space is used to indent lines in the final YAML file.
#Separators are used to separate enumerated values in nodes file.
###################################################################
space = ' '
separator = ','
######################################################################################
#The boolean and type maps convert values in the dictionary to YAML-compliant values.
######################################################################################
boolean_map= {'Yes': 'true', 'No': 'false'}
type_map = {'int': 'integer', 'float': 'number', 'string': 'string', 'integer': 'integer', 'boolean': 'boolean', 'number': 'number', 'datetime': 'datetime'}

#########################################################################################
#Dictionaries to hold properties of an attribute.
#Update this with a list of dictionaries every time a property is added to an attribute.
#########################################################################################
of_node = defaultdict(list)
of_property = defaultdict(list)
source_field = {}
of_type = {}
enumeration = defaultdict(list)
description = {}
is_required = {}
is_private = {}
minimum_constraint = {}
exclusiveMinimum_constraint = {}
maximum_constraint = {}
exclusiveMaximum_constraint = {}
displayName = {}
constraints = {}
data_type = {}
display = {}
has_units = {}
node_category = {}
node_definition = {}

##########################################################################################################
#Function to create the nodes and relationships file. Output file has the generic name: model_file.yaml. 
#Be sure to change it to a more appropriate file name.
#Function reads data_dict and lists each node and it's properties.
##########################################################################################################
def createModelFile(data_dict, defn_file = None):
	sheet_list = data_dict.sheetnames

	#read the definition file, if provided.
	if defn_file is not None:
		defn_file_sheet = defn_file.active
		for row in defn_file_sheet.iter_rows(min_row = 2):
				node_attributes = [str(cell.value).strip() for cell in row]
				node_category[node_attributes[0]] = node_attributes[1]
				node_definition[node_attributes[0]] = node_attributes[1]
		node_names = list(node_category.keys())
		#test if the nodes listed in the data dictionary and the definition file match. If not, return an error and exit.
		if node_names.sort() != sheet_list.sort():
		 	print("Node names in the Nodes file and Definition file do not match. Please check input data.")
		 	sys.exit()
	
	file = open("model_file.yaml", "w")
	file.write("Nodes:"+"\n")
	#sys.exit()
	#test = sheet_list[0:14]
	for sheet_name in sheet_list:
		data_dict.sheet = data_dict[sheet_name]
		file.write(2*space+sheet_name+":"+"\n")
		if sheet_name in node_category:
			file.write(4*space+"Category"+":"+space+node_category[sheet_name]+"\n")
		file.write(4*space+"Props"+":"+"\n")
		for row in data_dict.sheet.iter_rows(min_row = 2):
			model_entity_attributes = [str(cell.value).strip() for cell in row]
			if(model_entity_attributes[2] != 'None'):
				file.write(6*space+'-'+1*space+str(model_entity_attributes[2])+'\n')

#########################################################################################
#Gather the Node attributes. Store attribute properties in the appropriate dictionaries. 
#########################################################################################
def addNodeProps(data_dict):
	sheet_list = data_dict.sheetnames
	for sheet_name in sheet_list:
		data_dict.sheet = data_dict[sheet_name]
		for row in data_dict.sheet.iter_rows(min_row = 2):
			model_entity_attributes = [str(cell.value).strip() for cell in row]
			if(model_entity_attributes[2] != 'None'):
				#print(sheet_name + "\t" + model_entity_attributes[2]) # test code
				of_node[model_entity_attributes[2]].append(sheet_name)
				description[model_entity_attributes[2]] = model_entity_attributes[14]
				source_field[model_entity_attributes[2]] = model_entity_attributes[0]
				displayName[model_entity_attributes[2]] = model_entity_attributes[3]
				constraints[model_entity_attributes[2]] = model_entity_attributes[5]
				data_type[model_entity_attributes[2]] = type_map[model_entity_attributes[4]]
				
				if(model_entity_attributes[5] == 'Of Enumeration'):
					attribute_list = str(model_entity_attributes[10]).split(separator)
					attribute_list = [attribute.strip() for attribute in attribute_list]
					attribute_list = ["\""+attribute+"\"" for attribute in attribute_list]
					enumeration[model_entity_attributes[2]] = attribute_list
				elif(model_entity_attributes[15] != 'None'):
					has_units[model_entity_attributes[2]] = model_entity_attributes[15]
					of_type[model_entity_attributes[2]] = type_map[model_entity_attributes[4]]
				else:
					of_type[model_entity_attributes[2]] = type_map[model_entity_attributes[4]]
				
				is_required[model_entity_attributes[2]] = boolean_map[model_entity_attributes[11]]
				is_private[model_entity_attributes[2]] = boolean_map[model_entity_attributes[12]]
				display[model_entity_attributes[2]] = boolean_map[model_entity_attributes[13]]
				
				#Parse the numeric constraints, if any.
				if model_entity_attributes[6] != 'None':
					minimum_constraint[model_entity_attributes[2]] = model_entity_attributes[6]
				if model_entity_attributes[7] != 'None':
					exclusiveMinimum_constraint[model_entity_attributes[2]] = model_entity_attributes[7]
				if model_entity_attributes[8] != 'None':
					maximum_constraint[model_entity_attributes[2]] = model_entity_attributes[8]
				if model_entity_attributes[9] != 'None':
					exclusiveMaximum_constraint[model_entity_attributes[2]] = model_entity_attributes[9]

##########################################################################################################
#Read the Relationships listed in the "Edges" file. Print the relationships in the model.yaml file.
##########################################################################################################
def addRelationships(relations_dict):
	ends_dict = defaultdict(list)
	props_dict = defaultdict(list)
	mul_dict = {}
	sheet_list = relations_dict.sheetnames
	relationships_list = relations_dict[sheet_list[0]]
	for row in relationships_list.iter_rows(min_row = 2):
		relationship_attributes = [str(cell.value).strip() for cell in row]
		ends_dict[relationship_attributes[0]].append((relationship_attributes[1],relationship_attributes[2]))
		mul_dict[relationship_attributes[0]] = relationship_attributes[3]

	for sheet_name in sheet_list[1:]:
		relationship_sheet = relations_dict[sheet_name]
		for row in relationship_sheet.iter_rows(min_row = 2):
			relationship_attributes = [str(cell.value).strip() for cell in row]
			if(relationship_attributes[2] != 'None'):
				props_dict[sheet_name].append(relationship_attributes[2])
	
	file = open("model_file.yaml", "a")
	file.write('Relationships:'+'\n')
	for key in mul_dict.keys():
		file.write(2*space+key+':'+'\n')
		file.write(4*space+'Mul: '+ mul_dict[key]+'\n')
		file.write(4*space+'Ends:'+'\n')
		ends_list = ends_dict[key]
		for src, dst in ends_list:
			file.write(6*space+'-'+space+'Src: '+src+'\n')
			file.write(8*space+'Dst: '+dst+'\n')
		if(props_dict[key]):
			file.write(4*space+'Props:'+'\n')
			for prop in props_dict[key]:
				file.write(6*space+'-'+1*space+prop+'\n')
		else:
			file.write(4*space+'Props: null'+'\n')
#################################################################################################
#Gather the Relationship attributes. Store attribute properties in the appropriate dictionaries. 
#################################################################################################
def addRelationshipProps(relations_dict):
	sheet_list = relations_dict.sheetnames
	for sheet_name in sheet_list[1: ]:
		relations_dict.sheet = relations_dict[sheet_name]
		for row in relations_dict.sheet.iter_rows(min_row = 2):
			model_entity_attributes = [str(cell.value).strip() for cell in row]
			if(model_entity_attributes[2] != 'None'):
				of_property[model_entity_attributes[2]].append(sheet_name)
				description[model_entity_attributes[2]] = model_entity_attributes[14]
				source_field[model_entity_attributes[2]] = model_entity_attributes[0]
				displayName[model_entity_attributes[2]] = model_entity_attributes[3]
				constraints[model_entity_attributes[2]] = model_entity_attributes[5]
				data_type[model_entity_attributes[2]] = type_map[model_entity_attributes[4]]
				if(model_entity_attributes[5] == 'Of Enumeration'):
					attribute_list = str(model_entity_attributes[10]).split(separator)
					attribute_list = [attribute.strip() for attribute in attribute_list]
					attribute_list = ["\""+attribute+"\"" for attribute in attribute_list]
					enumeration[model_entity_attributes[2]] = attribute_list
				elif(model_entity_attributes[15] != 'None'):
					has_units[model_entity_attributes[2]] = model_entity_attributes[15]
					of_type[model_entity_attributes[2]] = type_map[model_entity_attributes[4]]
				else:
					of_type[model_entity_attributes[2]] = type_map[model_entity_attributes[4]]
				
				is_required[model_entity_attributes[2]] = boolean_map[model_entity_attributes[11]]
				is_private[model_entity_attributes[2]] = boolean_map[model_entity_attributes[12]]
				display[model_entity_attributes[2]] = boolean_map[model_entity_attributes[13]]
				
				#Parse the numeric constraints, if any.
				if model_entity_attributes[6] != 'None':
					minimum_constraint[model_entity_attributes[2]] = model_entity_attributes[6]
				if model_entity_attributes[7] != 'None':
					exclusiveMinimum_constraint[model_entity_attributes[2]] = model_entity_attributes[7]
				if model_entity_attributes[8] != 'None':
					maximum_constraint[model_entity_attributes[2]] = model_entity_attributes[8]
				if model_entity_attributes[9] != 'None':
					exclusiveMaximum_constraint[model_entity_attributes[2]] = model_entity_attributes[9]

########################################################################
#Print the unique list of properties in the model_properties_file.yaml.
########################################################################
def createPropsFile():
	file = open("model_properties_file.yaml", "w")
	file.write("PropDefinitions:"+"\n")
	for prop in of_node.keys():
		of_node_types = ", ".join(of_node[prop])
		file.write('#'+'property of '+of_node_types+'\n')
		file.write(2*space+prop+':'+'\n')
		file.write(4*space+'Desc: '+description[prop]+'\n')
		file.write(4*space+'Src: '+source_field[prop]+'\n')
		if(prop in of_type and prop not in has_units):
			file.write(4*space+'Type: '+of_type[prop]+'\n')
		elif(prop in of_type and prop in has_units):
			file.write(4*space+"Type:"+'\n')
			file.write(6*space+"units:"+'\n')
			file.write(7*space+'-'+1*space+has_units[prop]+'\n')
			file.write(6*space+"value_type: "+of_type[prop]+'\n')
		elif(prop in enumeration):
			file.write(4*space+"Type:"+'\n')
			for attribute in enumeration[prop]:
				file.write(6*space+'-'+1*space+attribute+'\n')
		
		file.write(4*space+'Req: '+is_required[prop]+'\n')
		file.write(4*space+'Private: '+is_private[prop]+'\n')
		
		#Print numeric constraints, if any.
		if(prop in minimum_constraint):
			file.write(4*space+'minimum: '+ minimum_constraint[prop]+'\n')
		if(prop in exclusiveMinimum_constraint):
			file.write(4*space+'exclusiveMinimum: '+ exclusiveMinimum_constraint[prop]+'\n')
		if(prop in maximum_constraint):
			file.write(4*space+'maximum: '+ maximum_constraint[prop]+'\n')
		if(prop in exclusiveMaximum_constraint):
			file.write(4*space+'exclusiveMaximum: '+ exclusiveMaximum_constraint[prop]+'\n')

	for prop in of_property.keys():
		of_property_types = ", ".join(of_property[prop])
		file.write('#'+'property of '+of_property_types+'\n')
		file.write(2*space+prop+':'+'\n')
		file.write(4*space+'Desc: '+description[prop]+'\n')
		file.write(4*space+'Src: '+source_field[prop]+'\n')
		if(prop in of_type and prop not in has_units):
			file.write(4*space+'Type: '+of_type[prop]+'\n')
		elif(prop in of_type and prop in has_units):
			file.write(4*space+"Type:"+'\n')
			file.write(6*space+"units:"+'\n')
			file.write(7*space+'-'+1*space+has_units[prop]+'\n')
			file.write(6*space+"value_type: "+of_type[prop]+'\n')
		elif(prop in enumeration):
			file.write(4*space+"Type:"+'\n')
			for attribute in enumeration[prop]:
				file.write(6*space+'-'+1*space+attribute+'\n')
		
		file.write(4*space+'Req: '+is_required[prop]+'\n')
		file.write(4*space+'Private: '+is_private[prop]+'\n')
		
		#Print numeric constraints, if any.
		if(prop in minimum_constraint):
			file.write(4*space+'minimum: '+ minimum_constraint[prop]+'\n')
		if(prop in exclusiveMinimum_constraint):
			file.write(4*space+'exclusiveMinimum: '+ exclusiveMinimum_constraint[prop]+'\n')
		if(prop in maximum_constraint):
			file.write(4*space+'maximum: '+ maximum_constraint[prop]+'\n')
		if(prop in exclusiveMaximum_constraint):
			file.write(4*space+'exclusiveMaximum: '+ exclusiveMaximum_constraint[prop]+'\n')

########################################################################
#Print the Data Dictionary.
########################################################################
def createDataDictionary():
	file = open("data_dictionary.txt", "w")
	for prop in of_node.keys():
		if is_private[prop] == 'false': #do not print properties that are private
			file.write("Attribute Name: " + prop +'\n')
			file.write("Description: " + description[prop] +'\n')
			of_node_types = ", ".join(of_node[prop])
			file.write("Attribute of Node: " + of_node_types +'\n')
			if (display[prop] == 'true'):
				file.write("Display Name: " + displayName[prop] +'\n') #print display name only if display flag for field is set to true.
			file.write("Required: " + is_required[prop] +'\n')
			file.write("Type: " + data_type[prop] +'\n')
			file.write("Constraints: " + constraints[prop] +'\n')
			if prop in enumeration:
				enumerated_values_for_prop = ", ".join(enumeration[prop])
			else:
				enumerated_values_for_prop = "None"
			file.write("Enumeration: " + enumerated_values_for_prop +'\n\n')
	
	for prop in of_property.keys():
		if is_private[prop] == 'false': #do not print properties that are private
			file.write("Attribute Name: " + prop +'\n')
			file.write("Description: " + description[prop] +'\n')
			of_property_types = ", ".join(of_property[prop])
			file.write("Attribute of Relationship: " + of_property_types +'\n')
			if (display[prop] == 'true'):
				file.write("Display Name: " + displayName[prop] +'\n') #print display name only if display flag for field is set to true.
			file.write("Required: " + is_required[prop] +'\n')
			file.write("Type: " + data_type[prop] +'\n')
			file.write("Constraints: " + constraints[prop] +'\n')
			if prop in enumeration:
				enumerated_values_for_prop = ", ".join(enumeration[prop])
			else:
				enumerated_values_for_prop = "None"
			file.write("Enumeration: " + enumerated_values_for_prop +'\n\n')

def main():
	parser = argparse.ArgumentParser(description='A script to create model description files in the YAML format.')
	parser.add_argument('-n', '--nodes', required = True, help = 'File that lists nodes and their properties.')
	parser.add_argument('-e', '--edges', required = True, help = 'File that lists edges and their properties.')
	parser.add_argument('-d', '--defn', required = False, help = 'File that stores node definitions and categories.')
	args = parser.parse_args()

	#Read the Nodes and Edges Files respectively.
	model_dictionary = load_workbook(filename = args.nodes, read_only = True)
	relationships_dictionary = load_workbook(filename = args.edges, read_only = True)
	if args.defn is not None:
		definitions_dictionary = load_workbook(filename = args.defn, read_only = True)
	else:
		definitions_dictionary = None
	
	createModelFile(model_dictionary, definitions_dictionary)
	addNodeProps(model_dictionary)
	addRelationships(relationships_dictionary)
	addRelationshipProps(relationships_dictionary)
	createPropsFile()
	createDataDictionary()
	print("Files model_file.yaml and model_properties_file.yaml and dictionary have been successfully created in current directory.")


if __name__ == '__main__':
	main()